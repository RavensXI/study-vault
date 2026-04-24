import json, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

specs = json.load(open('specs/index.json'))
status = json.load(open('data/free-tier-status.json'))

# ── Mapping: spec slug → supabase slug ──
SLUG_MAP = {
    'science-8464': 'science',
    'science-8465': None,  # AQA Synergy — not built
    'combined-science': 'science-edexcel',
    'combined-science-a-gateway': 'science-ocr',
    'combined-science-b-21c': None,  # OCR B — not built
    'mathematics-8300': 'maths-aqa',
    'mathematics': None,  # generic slug used by multiple boards — handled below
    'english-language-8700': 'english-language',
    'english-literature-8702': 'english-literature',
    'design-and-technology-8552': 'design-technology',
    'religious-studies-8061': None,
    'religious-studies-8062': 'religious-education',
    'religious-studies-8063': None,
    # Separate sciences — content now merged into Science subject
    'biology-8461': 'science',
    'chemistry-8462': 'science',
    'physics-8463': 'science',
}

NAME_MAP = {
    ('Edexcel', 'mathematics'): 'maths',
    ('OCR', 'mathematics'): 'maths-ocr',
    ('Eduqas', 'mathematics'): 'maths-eduqas',
    ('Edexcel', 'english language'): 'english-language-edexcel',
    ('OCR', 'english language'): 'english-language-ocr',
    ('Eduqas', 'english language'): 'english-language-eduqas',
    ('WJEC', 'english language'): 'english-language-eduqas',  # Eduqas is WJEC's English arm
    ('Edexcel', 'english literature'): 'english-literature-edexcel',
    ('OCR', 'english literature'): 'english-literature-ocr',
    ('Eduqas', 'english literature'): 'english-literature-eduqas',
    ('WJEC', 'english literature'): 'english-literature-eduqas',
    ('OCR', 'computer science'): 'computer-science',
    ('Edexcel', 'computer science'): None,  # not built — different spec from OCR
    ('Eduqas', 'computer science'): None,
    ('WJEC', 'computer science'): None,
    ('Edexcel', 'history'): 'history',
    ('Eduqas', 'history'): None,  # not built — different spec from Edexcel
    ('WJEC', 'history'): None,
    ('WJEC', 'level 1/2 vocational award in hospitality and catering'): 'hospitality-catering',
    ('Eduqas', 'level 1/2 vocational award in hospitality and catering'): 'hospitality-catering',
}

status_by_slug = {s['slug']: s for s in status}

def find_status(spec):
    slug = spec['slug']
    # Name map takes priority (handles ambiguous slugs like 'english-literature' shared across boards)
    key = (spec['board'], spec['subject'].lower())
    if key in NAME_MAP:
        if NAME_MAP[key] is None:
            return None  # explicitly marked as not built
        if NAME_MAP[key] in status_by_slug:
            return status_by_slug[NAME_MAP[key]]
    # Slug map (spec slug -> supabase slug)
    if slug in SLUG_MAP and SLUG_MAP[slug] and SLUG_MAP[slug] in status_by_slug:
        return status_by_slug[SLUG_MAP[slug]]
    # Direct match (only if unambiguous)
    if slug in status_by_slug:
        return status_by_slug[slug]
    return None

# ── Categories ──
CORE_NAMES = ['english language', 'english literature', 'mathematics',
              'combined science', 'biology', 'chemistry', 'physics']
HIGH_DEMAND = ['geography', 'history', 'french', 'spanish', 'german',
               'physical education', 'business', 'drama', 'music',
               'art and design', 'computer science', 'design and technology',
               'food preparation and nutrition', 'religious studies',
               'psychology', 'sociology', 'media studies',
               'italian', 'latin', 'statistics', 'citizenship',
               'dance', 'economics', 'film studies', 'engineering',
               'hospitality and catering', 'health and social care',
               'sport science', 'sport studies', 'child development',
               'enterprise and marketing', 'creative imedia']

def categorise(name):
    nl = name.lower()
    for c in CORE_NAMES:
        if c in nl:
            return 'Core'
    for h in HIGH_DEMAND:
        if h in nl:
            return 'High demand'
    return 'Other'

# ── Expected format per subject type ──
# Practice-first: entire subject is practice problems
PRACTICE_SUBJECTS = ['mathematics', 'english language', 'french', 'spanish', 'german',
                     'italian', 'mathematics and numeracy', 'latin', 'classical greek',
                     'arabic', 'bengali', 'chinese', 'greek', 'gujarati', 'hebrew',
                     'japanese', 'panjabi', 'persian', 'polish', 'portuguese', 'russian',
                     'turkish', 'urdu', 'welsh language', 'welsh second language',
                     'global business communication']
# Mixed: mostly article but with one or more practice units
MIXED_SUBJECTS = ['combined science', 'separate science', 'biology', 'chemistry', 'physics',
                  'geography', 'statistics', 'computer science', 'electronics',
                  'psychology', 'physical education', 'geology']
# Subjects with Foundation/Higher tiering
TIERED_SUBJECTS = ['mathematics', 'combined science', 'separate science', 'biology',
                   'chemistry', 'physics', 'french', 'spanish', 'german', 'italian',
                   'statistics', 'mathematics and numeracy']

def expected_format(name):
    nl = name.lower()
    for p in PRACTICE_SUBJECTS:
        if p in nl:
            return 'Practice'
    for m in MIXED_SUBJECTS:
        if m in nl:
            return 'Mixed'
    return 'Article'

# ── Build rows ──
rows = []
for spec in sorted(specs, key=lambda s: (
    0 if categorise(s['subject']) == 'Core' else 1 if categorise(s['subject']) == 'High demand' else 2,
    s['subject'].lower(), s['board']
)):
    st = find_status(spec)
    built = st is not None
    cat = categorise(spec['subject'])
    fmt = expected_format(spec['subject'])

    # Use expected format (from spec analysis), not DB format field
    actual_fmt = fmt

    # Practice data status
    if actual_fmt == 'Practice':
        if st and st.get('practice', '') and st['practice'] != 'N/A':
            practice_status = st['practice']
        elif built:
            practice_status = 'Needs checking'
        else:
            practice_status = 'Needs building'
    elif actual_fmt == 'Mixed':
        if built:
            # For Mixed subjects, check if practice_data exists in any lessons
            prac = st.get('practice', '') if st else ''
            if prac and prac != 'N/A' and prac != '':
                # Parse x/y — if x > 0, practice units have data
                parts = prac.split('/')
                try:
                    num = int(parts[0])
                    practice_status = prac if num > 0 else 'Needs building'
                except:
                    practice_status = prac
            else:
                practice_status = 'Needs building'
        else:
            practice_status = 'Needs building'
    else:
        practice_status = 'N/A'

    rows.append({
        'subject': spec['subject'],
        'board': spec['board'],
        'code': spec['spec_code'],
        'category': cat,
        'format': actual_fmt if built else fmt,
        'status': 'Built' if built else 'Not built',
        'lessons': st['lessons'] if st else '',
        'practice': practice_status,
        'heroes': st.get('heroes', '') if st else '',
        'narration': st.get('narration', '') if st else '',
        'podcasts': st.get('podcasts', '') if st else '',
        'relatedMedia': st.get('relatedMedia', '') if st else '',
        'kcs': st.get('knowledgeChecks', '') if st else '',
        'flashcards': st.get('flashcards', '') if st else '',
        'htWrapping': (st.get('htWrapping', '') if st else '') if any(t in spec['subject'].lower() for t in TIERED_SUBJECTS) else 'N/A',
    })

# ── Preservation: snapshot user-edited columns from existing xlsx ──
# Each run wipes the workbook and regenerates it. Before that happens, read the
# existing file and remember any user edits in the "Hero QA" / "Notes" / "QA
# Status" / "Priority" columns so we can restore them into the rebuilt sheets.
# Keyed by stable row identity so new/removed rows are handled cleanly.
import os as _os
OUT_PATH = 'data/gcse-subject-tracker.xlsx'
user_snapshot = {}  # { sheet_name: { row_key: { col_name: (value, fill_rgb or None) } } }
if _os.path.exists(OUT_PATH):
    try:
        existing = openpyxl.load_workbook(OUT_PATH)
        # Main tracker — key by (Subject, Board, Spec Code)
        main_sheet_name = 'GCSE Subject Tracker'
        if main_sheet_name in existing.sheetnames:
            ms = existing[main_sheet_name]
            hdr = [ms.cell(row=1, column=c).value for c in range(1, ms.max_column + 1)]
            try:
                cQA = hdr.index('QA Status') + 1
                cNotes = hdr.index('QA Notes') + 1
                cPri = hdr.index('Priority') + 1
            except ValueError:
                cQA = cNotes = cPri = None
            snap = {}
            for r in range(2, ms.max_row + 1):
                subj = ms.cell(row=r, column=1).value
                board = ms.cell(row=r, column=2).value
                code = ms.cell(row=r, column=3).value
                key = (subj, board, code)
                entry = {}
                for label, idx in [('QA Status', cQA), ('QA Notes', cNotes), ('Priority', cPri)]:
                    if not idx: continue
                    cell = ms.cell(row=r, column=idx)
                    fill_rgb = None
                    if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
                        rgb = cell.fill.fgColor.rgb
                        if rgb not in ('00000000', None):
                            fill_rgb = rgb
                    if cell.value or fill_rgb:
                        entry[label] = (cell.value, fill_rgb)
                if entry:
                    snap[key] = entry
            user_snapshot[main_sheet_name] = snap
        # Eng Lit Texts — key by Text / Unit column
        if 'English Lit Texts' in existing.sheetnames:
            ls = existing['English Lit Texts']
            hdr = [ls.cell(row=1, column=c).value for c in range(1, ls.max_column + 1)]
            user_cols = ['AQA Hero QA', 'Edexcel Hero QA', 'OCR Hero QA', 'Eduqas Hero QA', 'Notes']
            col_idx = {c: (hdr.index(c) + 1) for c in user_cols if c in hdr}
            snap = {}
            for r in range(2, ls.max_row + 1):
                key = ls.cell(row=r, column=1).value
                if not key: continue
                entry = {}
                for label, idx in col_idx.items():
                    cell = ls.cell(row=r, column=idx)
                    fill_rgb = None
                    if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
                        rgb = cell.fill.fgColor.rgb
                        if rgb not in ('00000000', None):
                            fill_rgb = rgb
                    if cell.value or fill_rgb:
                        entry[label] = (cell.value, fill_rgb)
                if entry:
                    snap[key] = entry
            user_snapshot['English Lit Texts'] = snap
        # Hero QA per-subject sheets — key by (Board, Unit, Lesson #)
        hero_qa_sheet_names = [
            'Combined Science', 'Separate Sciences', 'History', 'Religious Education',
            'Computer Science', 'Design & Technology', 'Health & Social Care',
            'Hospitality & Catering', 'Music Technology', 'Business Studies',
        ]
        for sn in hero_qa_sheet_names:
            if sn[:31] not in existing.sheetnames: continue
            hs = existing[sn[:31]]
            hdr = [hs.cell(row=1, column=c).value for c in range(1, hs.max_column + 1)]
            try:
                cHQA = hdr.index('Hero QA') + 1
                cN = hdr.index('Notes') + 1
            except ValueError:
                continue
            snap = {}
            for r in range(2, hs.max_row + 1):
                board = hs.cell(row=r, column=2).value
                unit = hs.cell(row=r, column=3).value
                lnum = hs.cell(row=r, column=4).value
                key = (board, unit, lnum)
                entry = {}
                for label, idx in [('Hero QA', cHQA), ('Notes', cN)]:
                    cell = hs.cell(row=r, column=idx)
                    fill_rgb = None
                    if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
                        rgb = cell.fill.fgColor.rgb
                        if rgb not in ('00000000', None):
                            fill_rgb = rgb
                    if cell.value or fill_rgb:
                        entry[label] = (cell.value, fill_rgb)
                if entry:
                    snap[key] = entry
            user_snapshot[sn[:31]] = snap
        _preserved_count = sum(len(v) for v in user_snapshot.values())
        if _preserved_count:
            print(f'[preservation] snapshotted {_preserved_count} user-edited rows across {len(user_snapshot)} sheets')
    except Exception as e:
        print(f'[preservation] WARNING: could not read existing xlsx: {e}')

# ── Create Excel ──
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'GCSE Subject Tracker'

hfont = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
hfill = PatternFill(start_color='2D2A26', end_color='2D2A26', fill_type='solid')
built_f = PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid')
notbuilt_f = PatternFill(start_color='FFCDD2', end_color='FFCDD2', fill_type='solid')
core_f = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')
high_f = PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid')
partial_f = PatternFill(start_color='FFF9C4', end_color='FFF9C4', fill_type='solid')
na_font = Font(color='AAAAAA', italic=True)
practice_f = PatternFill(start_color='E8EAF6', end_color='E8EAF6', fill_type='solid')
mixed_f = PatternFill(start_color='E0F2F1', end_color='E0F2F1', fill_type='solid')
article_f = PatternFill(start_color='FFF8E1', end_color='FFF8E1', fill_type='solid')
needs_f = PatternFill(start_color='FFCDD2', end_color='FFCDD2', fill_type='solid')
tb = Border(left=Side(style='thin', color='D6D3D1'), right=Side(style='thin', color='D6D3D1'),
            top=Side(style='thin', color='D6D3D1'), bottom=Side(style='thin', color='D6D3D1'))

headers = ['Subject', 'Exam Board', 'Spec Code', 'Category', 'Format', 'Status',
           'Lessons', 'Practice Data', 'Heroes', 'Narration', 'Podcasts',
           'Related Media', 'Knowledge Checks', 'Flashcards', 'F/H Wrapping',
           'QA Status', 'QA Notes', 'Priority']

for ci, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=ci, value=h)
    cell.font = hfont
    cell.fill = hfill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = tb

for ri, r in enumerate(rows, 2):
    vals = [r['subject'], r['board'], r['code'], r['category'], r['format'], r['status'],
            r['lessons'], r['practice'], r['heroes'], r['narration'], r['podcasts'],
            r['relatedMedia'], r['kcs'], r['flashcards'], r['htWrapping'], '', '', '']
    for ci, v in enumerate(vals, 1):
        cell = ws.cell(row=ri, column=ci, value=v)
        cell.border = tb
        cell.alignment = Alignment(horizontal='center', vertical='center')

        # Category colour
        if ci == 4:
            if v == 'Core': cell.fill = core_f
            elif v == 'High demand': cell.fill = high_f

        # Format colour
        if ci == 5:
            if v == 'Practice': cell.fill = practice_f
            elif v == 'Mixed': cell.fill = mixed_f
            elif v == 'Article': cell.fill = article_f

        # Status colour
        if ci == 6:
            cell.fill = built_f if v == 'Built' else notbuilt_f

        # Practice data column
        if ci == 8:
            if v == 'N/A':
                cell.font = na_font
            elif v == 'Needs building':
                cell.fill = notbuilt_f
            elif v == 'Needs checking':
                cell.fill = partial_f
            elif isinstance(v, str) and '/' in v:
                parts = v.split('/')
                try:
                    num, den = int(parts[0]), int(parts[1])
                    if num == den and den > 0: cell.fill = built_f
                    elif num == 0: cell.fill = notbuilt_f
                    elif num < den: cell.fill = partial_f
                except: pass

        # Completion columns (heroes, narration, etc.)
        if isinstance(v, str) and '/' in v and ci >= 9 and ci <= 15:
            parts = v.split('/')
            try:
                num, den = int(parts[0]), int(parts[1])
                if num == den and den > 0: cell.fill = built_f
                elif num == 0: cell.fill = notbuilt_f
                elif num < den: cell.fill = partial_f
            except: pass

widths = [40, 12, 10, 13, 10, 10, 8, 14, 10, 10, 10, 13, 15, 11, 14, 12, 25, 10]
for ci, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(ci)].width = w

ws.freeze_panes = 'A2'
ws.auto_filter.ref = ws.dimensions

# ── Supabase client for supplementary sheets ──
import os, sys
sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__))))
from lib.supabase_client import get_client
sb = get_client()

# ── English Literature Texts sheet ──
# Cross-board matrix of every text, with per-board hero QA columns
lit_ws = wb.create_sheet(title='English Lit Texts')

lit_headers = ['Text / Unit', 'AQA', 'Edexcel', 'OCR', 'Eduqas', 'Boards', 'Overlap',
               'AQA Hero QA', 'Edexcel Hero QA', 'OCR Hero QA', 'Eduqas Hero QA', 'Notes']
for ci, h in enumerate(lit_headers, 1):
    cell = lit_ws.cell(row=1, column=ci, value=h)
    cell.font = hfont
    cell.fill = hfill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = tb

BOARD_SLUGS = {
    'AQA': 'english-literature',
    'Edexcel': 'english-literature-edexcel',
    'OCR': 'english-literature-ocr',
    'Eduqas': 'english-literature-eduqas',
}

# Build text → {board: unit_name} map
texts = {}  # normalised_name -> {AQA: 'Unit Name', Edexcel: ..., display: 'Unit Name'}

def normalise_title(name):
    """Normalise text names so variants like 'Jekyll & Hyde' and 'Jekyll and Hyde' match."""
    n = name.lower()
    n = n.replace('&', 'and').replace('dr ', 'dr-')
    n = ''.join(c if c.isalnum() or c in ' -' else '' for c in n).strip()
    n = ' '.join(n.split())
    return n

for board, slug in BOARD_SLUGS.items():
    res = sb.from_('subjects').select('id').eq('slug', slug).is_('school_id', 'null').execute()
    if not res.data:
        continue
    subj_id = res.data[0]['id']
    units = sb.from_('units').select('name, slug').eq('subject_id', subj_id).order('sort_order').execute()
    for u in (units.data or []):
        key = normalise_title(u['name'])
        if key not in texts:
            texts[key] = {'display': u['name']}
        texts[key][board] = u['name']

# Sort by board count desc, then by name
text_rows = sorted(texts.values(),
                   key=lambda t: (-sum(1 for b in ['AQA', 'Edexcel', 'OCR', 'Eduqas'] if b in t), t['display']))

for ri, t in enumerate(text_rows, 2):
    boards_count = sum(1 for b in ['AQA', 'Edexcel', 'OCR', 'Eduqas'] if b in t)
    if boards_count == 4:
        overlap = 'All 4 boards'
    elif boards_count == 1:
        overlap = '1 board only'
    else:
        overlap = f'{boards_count} boards'
    vals = [
        t['display'],
        'Yes' if 'AQA' in t else None,
        'Yes' if 'Edexcel' in t else None,
        'Yes' if 'OCR' in t else None,
        'Yes' if 'Eduqas' in t else None,
        boards_count,
        overlap,
        '', '', '', '', '',  # Hero QA columns + Notes, blank for Tom
    ]
    for ci, v in enumerate(vals, 1):
        cell = lit_ws.cell(row=ri, column=ci, value=v)
        cell.border = tb
        cell.alignment = Alignment(horizontal='center' if ci != 1 else 'left', vertical='center')
        # Colour cells: present=green, absent=light grey
        if ci in (2, 3, 4, 5):
            if v == 'Yes':
                cell.fill = built_f
            else:
                cell.fill = PatternFill(start_color='F5F1EC', end_color='F5F1EC', fill_type='solid')
                cell.font = na_font

lit_widths = [38, 8, 10, 8, 10, 9, 14, 14, 16, 14, 14, 30]
for ci, w in enumerate(lit_widths, 1):
    lit_ws.column_dimensions[get_column_letter(ci)].width = w
lit_ws.freeze_panes = 'A2'
lit_ws.auto_filter.ref = lit_ws.dimensions

# ── Hero QA sheets per subject ──
# Subjects that need hero image QA (non-practice, free tier)

PRACTICE_SLUGS = ['maths', 'maths-aqa', 'maths-ocr', 'maths-eduqas',
                  'english-language', 'english-language-edexcel', 'english-language-ocr', 'english-language-eduqas',
                  'spanish', 'french', 'german']

# English Lit handled separately (already has Eng Lit Texts sheet)
ENG_LIT_SLUGS = ['english-literature', 'english-literature-edexcel', 'english-literature-ocr', 'english-literature-eduqas']

# Group subjects into sheets
SHEET_GROUPS = [
    ('Combined Science', ['science', 'science-edexcel', 'science-ocr']),
    ('Separate Sciences', ['separate-sciences']),
    ('History', ['history']),
    ('Religious Education', ['religious-education']),
    ('Computer Science', ['computer-science']),
    ('Design & Technology', ['design-technology']),
    ('Health & Social Care', ['health-social-care']),
    ('Hospitality & Catering', ['hospitality-catering']),
    ('Music Technology', ['music-technology']),
    ('Business Studies', ['business-aqa', 'business-edexcel', 'business-ocr']),
]

hero_qa_total = 0

for sheet_name, slugs in SHEET_GROUPS:
    qa_ws = wb.create_sheet(title=sheet_name[:31])  # Excel 31-char limit

    qa_headers = ['Subject', 'Board', 'Unit', 'Lesson #', 'Lesson Title', 'Has Hero', 'Hero QA', 'Notes']
    for ci, h in enumerate(qa_headers, 1):
        cell = qa_ws.cell(row=1, column=ci, value=h)
        cell.font = hfont
        cell.fill = hfill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = tb

    qa_row = 2
    for slug in slugs:
        subj_res = sb.from_('subjects').select('id, slug, name, settings').eq('slug', slug).is_('school_id', 'null').execute()
        if not subj_res.data:
            continue
        subj = subj_res.data[0]
        settings = subj.get('settings') or {}
        if isinstance(settings, str):
            import json as _json
            settings = _json.loads(settings)
        practice_units = settings.get('practice_units', [])

        # Determine board from slug or name
        board = ''
        if 'edexcel' in slug: board = 'Edexcel'
        elif 'ocr' in slug: board = 'OCR'
        elif 'eduqas' in slug: board = 'Eduqas'
        else: board = 'AQA'

        units_res = sb.from_('units').select('id, name, slug').eq('subject_id', subj['id']).order('sort_order').execute()
        for unit in (units_res.data or []):
            if unit['slug'] in practice_units:
                continue
            lessons_res = sb.from_('lessons').select('id, lesson_number, title, hero_image_url').eq('unit_id', unit['id']).order('lesson_number').execute()
            for lesson in (lessons_res.data or []):
                has_hero = 'Yes' if lesson.get('hero_image_url') else 'No'
                vals = [subj['name'], board, unit['name'], lesson['lesson_number'], lesson['title'], has_hero, '', '']
                for ci, v in enumerate(vals, 1):
                    cell = qa_ws.cell(row=qa_row, column=ci, value=v)
                    cell.border = tb
                    cell.alignment = Alignment(horizontal='center' if ci != 5 else 'left', vertical='center')
                    if ci == 6:
                        cell.fill = built_f if v == 'Yes' else notbuilt_f
                qa_row += 1
                hero_qa_total += 1

    qa_widths = [30, 10, 30, 9, 45, 10, 10, 25]
    for ci, w in enumerate(qa_widths, 1):
        qa_ws.column_dimensions[get_column_letter(ci)].width = w
    qa_ws.freeze_panes = 'A2'
    qa_ws.auto_filter.ref = qa_ws.dimensions

# ── Preservation: restore user-edited columns from snapshot ──
# For each sheet we rebuilt, look up row keys and write preserved values
# (and fill colour) back into the user-edit columns.
_restored = 0

def _restore_fill(cell, rgb):
    if not rgb: return
    try:
        cell.fill = PatternFill(start_color=rgb, end_color=rgb, fill_type='solid')
    except Exception:
        pass

# Main tracker sheet
if 'GCSE Subject Tracker' in user_snapshot:
    snap = user_snapshot['GCSE Subject Tracker']
    hdr = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    col_idx = {c: (hdr.index(c) + 1) for c in ['QA Status', 'QA Notes', 'Priority'] if c in hdr}
    for r in range(2, ws.max_row + 1):
        key = (ws.cell(row=r, column=1).value, ws.cell(row=r, column=2).value, ws.cell(row=r, column=3).value)
        if key not in snap: continue
        for label, (val, rgb) in snap[key].items():
            if label not in col_idx: continue
            cell = ws.cell(row=r, column=col_idx[label])
            if val is not None: cell.value = val
            _restore_fill(cell, rgb)
            _restored += 1

# Eng Lit Texts
if 'English Lit Texts' in user_snapshot and 'English Lit Texts' in wb.sheetnames:
    ls = wb['English Lit Texts']
    snap = user_snapshot['English Lit Texts']
    hdr = [ls.cell(row=1, column=c).value for c in range(1, ls.max_column + 1)]
    col_idx = {c: (hdr.index(c) + 1) for c in ['AQA Hero QA', 'Edexcel Hero QA', 'OCR Hero QA', 'Eduqas Hero QA', 'Notes'] if c in hdr}
    for r in range(2, ls.max_row + 1):
        key = ls.cell(row=r, column=1).value
        if key not in snap: continue
        for label, (val, rgb) in snap[key].items():
            if label not in col_idx: continue
            cell = ls.cell(row=r, column=col_idx[label])
            if val is not None: cell.value = val
            _restore_fill(cell, rgb)
            _restored += 1

# Hero QA per-subject sheets
for sn_full, _ in SHEET_GROUPS:
    sn = sn_full[:31]
    if sn not in user_snapshot or sn not in wb.sheetnames: continue
    hs = wb[sn]
    snap = user_snapshot[sn]
    hdr = [hs.cell(row=1, column=c).value for c in range(1, hs.max_column + 1)]
    col_idx = {c: (hdr.index(c) + 1) for c in ['Hero QA', 'Notes'] if c in hdr}
    for r in range(2, hs.max_row + 1):
        key = (hs.cell(row=r, column=2).value, hs.cell(row=r, column=3).value, hs.cell(row=r, column=4).value)
        if key not in snap: continue
        for label, (val, rgb) in snap[key].items():
            if label not in col_idx: continue
            cell = hs.cell(row=r, column=col_idx[label])
            if val is not None: cell.value = val
            _restore_fill(cell, rgb)
            _restored += 1

if _restored:
    print(f'[preservation] restored {_restored} user-edited cells')

wb.save('data/gcse-subject-tracker.xlsx')

bc = sum(1 for r in rows if r['status'] == 'Built')
print(f'Written: {len(rows)} specs, {bc} built, {len(rows)-bc} not built')
print(f"Core: {sum(1 for r in rows if r['category']=='Core')} ({sum(1 for r in rows if r['category']=='Core' and r['status']=='Built')} built)")
print(f"High demand: {sum(1 for r in rows if r['category']=='High demand')} ({sum(1 for r in rows if r['category']=='High demand' and r['status']=='Built')} built)")
print(f"Other: {sum(1 for r in rows if r['category']=='Other')} ({sum(1 for r in rows if r['category']=='Other' and r['status']=='Built')} built)")
print(f"Hero QA sheets: {len(SHEET_GROUPS)} sheets, {hero_qa_total} lessons")
print()
print("Practice subjects not yet built:")
for r in rows:
    if r['format'] == 'Practice' and r['status'] == 'Not built':
        print(f"  {r['subject']} ({r['board']}) — {r['practice']}")
